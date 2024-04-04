# -*- encoding: utf-8 -*-
"""
Copyright (c) 2021 - present Orange Cyberdefense
"""

import json
import os
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from zipfile import BadZipFile, ZipFile

from flask import (
    current_app,
    flash,
    redirect,
    render_template,
    send_file,
    url_for,
    request,
)
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename

from app import db
from app.constants import EXTRACT_FOLDER_NAME, LANGUAGES_DEVICONS, PROJECTS_SRC_PATH
from app.base import util
from app.projects import blueprint
from app.projects.forms import ProjectForm, ExcelForm
from app.projects.models import Project
from app.projects.util import (
    check_zipfile,
    count_lines,
    has_access,
    remove_project,
    sha256sum,
    top_supported_language_lines_counts,
    get_user_projects_ids,
)
from app.base.models import Team


@blueprint.route("/projects")
@login_required
def projects_list():
    projects = Project.query.all()
    project_form = ProjectForm()
    admin = util.is_admin(current_user.role)
    user_projects_ids = get_user_projects_ids(current_user)
    return render_template(
        "projects_list.html",
        projects=projects,
        form=project_form,
        user=current_user,
        admin=admin,
        user_projects_ids=user_projects_ids,
        top_supported_language_lines_counts=top_supported_language_lines_counts,
        lang_icons=LANGUAGES_DEVICONS,
        segment="projects",
    )


@blueprint.route("/projects/<project_id>", methods=["POST", "GET"])
@login_required
def projects_dashboard(project_id):
    project = Project.query.filter_by(id=project_id).first_or_404()
    # Check if the user has access to the project
    if not has_access(current_user, project):
        return render_template("403.html"), 403
    # XLS export button
    form = ExcelForm()
    if form.validate_on_submit():
        selected_option = form.choice.data
        # Faire quelque chose avec l'option sélectionnée, comme rediriger vers une autre page
        return redirect(
            url_for(
                "projects_blueprint.scan_to_excel",
                project_id=project_id,
                selected_option=selected_option,
            )
        )
    # Get the 5 first inspector matches
    features = project.appinspector.match[0:5]
    return render_template(
        "project_dashboard.html",
        project=project,
        user=current_user,
        top_supported_language_lines_counts=top_supported_language_lines_counts,
        lang_icons=LANGUAGES_DEVICONS,
        features=features,
        segment="projects",
        form=form,
    )


@blueprint.route("/projects/<project_id>/status")
@login_required
def projects_status(project_id):
    project = Project.query.filter_by(id=project_id).first_or_404()
    # Check if the user has access to the project
    if not has_access(current_user, project):
        return render_template("403.html"), 403
    return str(project.status), 200


@blueprint.route("/projects/remove/<project_id>")
@login_required
def projects_remove(project_id):
    project = Project.query.filter_by(id=project_id).first_or_404()
    # Check if the user has access to the project
    if not has_access(current_user, project):
        return render_template("403.html"), 403
    remove_project(project)
    # If there is project excel generated, delet it
    path = os.getcwd()
    print(path)
    if os.path.exists(f"excel_save/{project.name}.xlsx"):
        os.remove(f"excel_save/{project.name}.xlsx")
    current_app.logger.info("Project deleted (project.id=%i)", project.id)
    flash("Project successfully deleted", "success")
    return redirect(url_for("projects_blueprint.projects_list"))


@blueprint.route("/projects/create", methods=["POST"])
@login_required
def projects_create():
    project_form = ProjectForm()
    # Form is valid
    if project_form.validate_on_submit():
        # Create a new project
        project_name = project_form.name.data
        file = project_form.source_archive.data
        project = Project(
            name=project_name,
            creator=current_user,
            archive_filename=secure_filename(file.filename),
        )
        # Store the new project in db
        db.session.add(project)
        db.session.commit()
        # Store the archive on disk
        if not os.path.isdir(PROJECTS_SRC_PATH):
            pathlib.Path(PROJECTS_SRC_PATH).mkdir(parents=True, exist_ok=True)
        project_path = os.path.join(PROJECTS_SRC_PATH, str(project.id))
        os.mkdir(project_path)
        archive_path = os.path.join(project_path, secure_filename(file.filename))
        file.save(archive_path)
        # Check if the provided archive is valid
        error, msg = check_zipfile(archive_path)
        if error:
            current_app.logger.warning(
                "Invalid archive file uploaded (archive path was '%s')", archive_path
            )
            remove_project(project)
            return msg, 403
        # Extract archive on disk
        project.archive_sha256sum = sha256sum(archive_path)
        source_path = os.path.join(project_path, EXTRACT_FOLDER_NAME)
        with ZipFile(archive_path, "r") as zip_ref:
            try:
                zip_ref.extractall(source_path)
            except BadZipFile as e:
                current_app.logger.error("Bad zip file: %s", str(e))
                remove_project(project)
                return "Bad zip file", 403
        # Count lines of code and save project
        count_lines(project)
        db.session.commit()
        current_app.logger.info("New project created (project.id=%i)", project.id)
        # Add the project to the global team
        Global_team = Team.query.filter_by(name="Global").first()
        Global_team_project_ids = [project.id for project in Global_team.projects]
        Global_team_project_ids.append(project.id)
        Global_team.projects = Project.query.filter(
            Project.id.in_(Global_team_project_ids)
        ).all()
        print(Global_team)
        db.session.commit()
        return str(project.id), 200
    # Form is invalid
    else:
        current_app.logger.warning(
            "Project add form invalid entries: %s", json.dumps(project_form.errors)
        )
        return json.dumps(project_form.errors), 403


@blueprint.route("/projects/to_excel/<project_id>")
@login_required
def scan_to_excel(project_id):
    project = Project.query.filter_by(id=project_id).first_or_404()
    # Check if the user has access to the project
    if not has_access(current_user, project):
        return render_template("403.html"), 403
    wb = Workbook()
    wb.is_spellcheck_enabled = False
    vulnerabilities = project.analysis.vulnerabilities
    selected_option = request.args.get("selected_option")

    # Create the sheets table
    sheets = {}
    # Set the first sheet
    sheets["MainSheet"] = wb.active
    sheets["MainSheet"].title = "MainSheet"
    # Set all vulnerabilitie sheets
    for index, vulnerabilitie in enumerate(vulnerabilities):
        sheet_name = f"vulnerabilitie {vulnerabilitie.id}"
        sheets[sheet_name] = wb.create_sheet(title=sheet_name)

    # Set the MainSheet data
    sheets["MainSheet"].merge_cells("A1:B1")
    sheets["MainSheet"].merge_cells("A2:B2")
    sheets["MainSheet"]["A1"] = "Project name"
    sheets["MainSheet"]["A2"] = project.name
    sheets["MainSheet"]["C1"] = "Project id"
    sheets["MainSheet"]["C2"] = project.id
    sheets["MainSheet"]["D1"] = "Risk level"
    sheets["MainSheet"]["D2"] = project.risk_level
    sheets["MainSheet"]["E1"] = "Lines"
    sheets["MainSheet"]["E2"] = project.project_lines_count.total_line_count

    # Set the vulnerabilites header
    sheets["MainSheet"].row_dimensions[4].height = 40
    sheets["MainSheet"].merge_cells("A4:C4")
    sheets["MainSheet"].merge_cells("G4:J4")
    sheets["MainSheet"].merge_cells("M4:V4")
    for row in sheets["MainSheet"]["A4":"X4"]:
        for cell in row:
            # Center text
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Set cells background color
            cell.fill = PatternFill(
                start_color="DEE6BD", end_color="DEE6BD", fill_type="solid"
            )
    sheets["MainSheet"]["A4"] = "Title"
    sheets["MainSheet"]["D4"] = "Id"
    sheets["MainSheet"]["E4"] = "Confidence"
    sheets["MainSheet"].column_dimensions["E"].width = 15
    sheets["MainSheet"]["F4"] = "Severity"
    sheets["MainSheet"]["G4"] = "Owasp"
    sheets["MainSheet"]["K4"] = "Impact"
    sheets["MainSheet"]["L4"] = "Likelihood"
    sheets["MainSheet"].column_dimensions["L"].width = 15
    sheets["MainSheet"]["M4"] = "Description"
    sheets["MainSheet"]["W4"] = "Occurences"
    sheets["MainSheet"].column_dimensions["W"].width = 15
    sheets["MainSheet"]["X4"] = "Ctrl Click"
    # Take all vulnerabilities
    for index, vulnerabilitie in enumerate(vulnerabilities):
        # Upgrade de rosheets["MainSheet"] size
        sheets["MainSheet"].row_dimensions[5 + index].height = 40
        # Set auto ligne return for description
        sheets["MainSheet"][f"M{5 + index}"].alignment = Alignment(wrap_text=True)
        # Set vulnerabilitie data
        sheets["MainSheet"].merge_cells(f"A{5 + index}:C{5 + index}")
        sheets["MainSheet"].merge_cells(f"G{5 + index}:J{5 + index}")
        sheets["MainSheet"].merge_cells(f"M{5 + index}:V{5 + index}")
        sheets["MainSheet"][f"A{5 + index}"] = vulnerabilitie.title
        sheets["MainSheet"][f"D{5 + index}"] = vulnerabilitie.id
        sheets["MainSheet"][f"E{5 + index}"] = vulnerabilitie.confidence
        sheets["MainSheet"][f"F{5 + index}"] = vulnerabilitie.severity
        sheets["MainSheet"][f"G{5 + index}"] = vulnerabilitie.owasp
        sheets["MainSheet"][f"K{5 + index}"] = vulnerabilitie.impact
        sheets["MainSheet"][f"L{5 + index}"] = vulnerabilitie.likelihood
        sheets["MainSheet"][f"M{5 + index}"] = vulnerabilitie.description
        sheets["MainSheet"][f"W{5 + index}"] = len(vulnerabilitie.occurences)
        # Ceate the vulnerabilitie occurences sheets button
        sheets["MainSheet"][f"X{5 + index}"] = f"{vulnerabilitie.id} details"
        sheets["MainSheet"][
            f"X{5 + index}"
        ].hyperlink = f"#'vulnerabilitie {vulnerabilitie.id}'!A1"

    # Set the vulnerabilitie occurences sheets
    for index, vulnerabilitie in enumerate(vulnerabilities):
        # Set the mainSheet button
        sheet_name = f"vulnerabilitie {vulnerabilitie.id}"
        sheets[sheet_name].column_dimensions["A"].width = 15
        sheets[sheet_name]["A1"] = "MainSheet"
        # Set sheet data
        sheets[sheet_name]["A1"].hyperlink = f"#'MainSheet'!A1"
        sheets[sheet_name].column_dimensions["B"].width = 15
        sheets[sheet_name]["B1"] = f"Vulnerabilitie {vulnerabilitie.id}"
        sheets[sheet_name].merge_cells("C1:D1")
        sheets[sheet_name].column_dimensions["C"].width = 15
        sheets[sheet_name]["C1"] = vulnerabilitie.title

        # Set occurences header
        sheets[sheet_name].merge_cells("A3:D3")
        sheets[sheet_name]["A3"] = "File path"
        sheets[sheet_name]["E3"] = "Id"
        sheets[sheet_name].merge_cells("F3:Z3")
        sheets[sheet_name]["F3"] = "Match_string"

        for row in sheets[sheet_name]["A3":"Z3"]:
            for cell in row:
                # Center header text
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Set cells background color
                cell.fill = PatternFill(
                    start_color="DEE6BD", end_color="DEE6BD", fill_type="solid"
                )

        # Set occurences data
        for index, occurence in enumerate(vulnerabilitie.occurences):
            if (
                (selected_option == "only_confirmed" and occurence.status == 1)
                or (
                    selected_option == "confirmed_and_undefined"
                    and occurence.status in [1, 0]
                )
                or (selected_option == "all")
            ):
                sheets[sheet_name].merge_cells(f"A{4 + index}:D{4 + index}")
                sheets[sheet_name][f"A{4 + index}"] = occurence.file_path
                sheets[sheet_name][f"E{4 + index}"] = occurence.id
                sheets[sheet_name].merge_cells(f"F{4 + index}:Z{4 + index}")
                sheets[sheet_name][f"F{4 + index}"] = occurence.match_string

    wb.save(f"excel_save/{project.name}.xlsx")
    # The excel file is deleted along with the proejct in /project/remove/<project_id>
    return send_file(f"../excel_save/{project.name}.xlsx", as_attachment=True)
