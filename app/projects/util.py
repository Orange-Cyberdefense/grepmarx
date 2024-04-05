# -*- encoding: utf-8 -*-
"""
Copyright (c) 2021 - present Orange Cyberdefense
"""

import json
import os
import pathlib
import subprocess
from hashlib import sha256
from shutil import rmtree
from zipfile import ZipFile, is_zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

from app import db
from app.constants import (
    EXPORT_FOLDER_NAME,
    EXTRACT_FOLDER_NAME,
    PROJECTS_SRC_PATH,
    ROLE_ADMIN,
    SCC,
    SEVERITY_CRITICAL,
    SEVERITY_HIGH,
    SEVERITY_LOW,
    SEVERITY_MEDIUM,
)
from app.projects.models import LanguageLinesCount, ProjectLinesCount
from app.rules.models import SupportedLanguage
from app.base.models import Team

##
## Project utils
##


def remove_project(project):
    """Delete the project from the database (along with all its analysis), and
    remove the project folder from disk.

    Args:
        project (Project): project to remove
    """
    project_path = os.path.join(PROJECTS_SRC_PATH, str(project.id))
    if os.path.isdir(project_path):
        rmtree(project_path)
    db.session.delete(project)
    db.session.commit()


def count_lines(project):
    """Count line of code of the project's code archive using third-party tool
    scc, and populate the ProjectLinesCount class member.

    Args:
        project (project): project with an already extracted source archive
    """
    source_path = os.path.join(PROJECTS_SRC_PATH, str(project.id), EXTRACT_FOLDER_NAME)
    # Call to external binary: scc
    json_result = json.loads(
        subprocess.run([SCC, source_path, "-f", "json"], capture_output=True).stdout
    )
    project.project_lines_count = load_project_lines_count(json_result)


def sha256sum(file_path):
    """Calculate the SHA256 sum of a file.

    Args:
        file_path (str): file for which to calculate the sum

    Returns:
        str: digest of the SHA256 sum
    """
    sha256_hash = sha256()
    with open(file_path, "rb") as f:
        # Read and update hash string value in blocks of 4K
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()


def check_zipfile(zip_path):
    """Check if a zip file is valid and unencrypted.

    Args:
        zip_path (str): path to the zip file

    Returns:
        [bool]: true if the file is valid and unencrypted
        [str]: short error message if the file is invalid or encrypted
    """
    error = False
    msg = ""
    if not is_zipfile(zip_path):
        error = True
        msg = "invalid zip file"
    else:
        for zinfo in ZipFile(zip_path, "r").infolist():
            if zinfo.flag_bits & 0x1:
                error = True
                msg = "encrypted zip file"
                break
    return error, msg


def count_occurences(project):
    """Count the total number of vulnerability occurences for a project.
    An analysis must be associated to this project.

    Args:
        project (Project): Project for which occurences must be counted

    Returns:
        int: Total vulnerability occurences for the project
    """
    occurences_count = 0
    # We need an analysis to count occurences
    if project.analysis is not None:
        # Count the total number of vulnerability occurences
        for c_vulnerability in project.analysis.vulnerabilities:
            occurences_count = occurences_count + len(c_vulnerability.occurences)
    return occurences_count


def calculate_risk_level(project):
    """Calculate the risk level for a project. An analysis must be associated
    to this project. The risk level is calculated mainly from the severity level of the
    found vulnerabilities (SAST), and adjusted with the severity of it's dependencies'
    known vulnerabilities (SCA).

    Args:
        project (Project): Project for which the risk level has to be calculated

    Returns:
        int: Calculated risk level (0 - 100)
    """
    risk_level = 0
    # We need an analysis to calculate a risk level
    if project.analysis is not None:
        # Make sure LOC count is > 0
        if (
            project.project_lines_count is not None
            and project.project_lines_count.total_code_count > 0
        ):
            # 1. Define a base risk level depending on the vulns' severities
            s2rl = {
                SEVERITY_CRITICAL: 75,
                SEVERITY_HIGH: 60,
                SEVERITY_MEDIUM: 40,
                SEVERITY_LOW: 20,
            }
            for s in s2rl:
                if has_vuln_with_severity(project.analysis, s):
                    risk_level = s2rl[s]
                    break
            # 2. Adjust the level with SCA results' severities
            s2rl = {
                SEVERITY_CRITICAL: 10,
                SEVERITY_HIGH: 8,
                SEVERITY_MEDIUM: 5,
                SEVERITY_LOW: 2,
            }
            for s in s2rl:
                if has_vuln_dep_with_severity(project.analysis, s):
                    risk_level += s2rl[s]
    return risk_level


def has_vuln_with_severity(analysis, severity_level):
    """Check if an analysis contains at least one vulnerability with
    the given severity level.

    Args:
        analysis (Analysis): analysis populated with vulnerabilities
        severity_level (str): severity level to search in analysis' vulnerabilities

    Returns:
        bool: True if the analysis contains at least one vulnerability with
        the given severity level
    """
    for vuln in analysis.vulnerabilities:
        if vuln.severity == severity_level:
            return True
    return False


def has_vuln_dep_with_severity(analysis, severity_level):
    """Check if an analysis contains at least one vulnerable dependency with
    the given severity level.

    Args:
        analysis (Analysis): analysis populated with vulnerable dependencies
        severity_level (str): severity level to search in analysis' vulnerable dependencies

    Returns:
        bool: True if the analysis contains at least one vulnerable dependency with
        the given severity level
    """
    for vuln in analysis.vulnerable_dependencies:
        if vuln.severity == severity_level:
            return True
    return False


##
## ProjectLinesCount util
##


def top_language_lines_counts(project_lc, top_number):
    """Return the `top_number` most present languages in the project source archive,
    sorted by their lines counts.

    Args:
        project_lc (ProjectLinesCount): project lines count object populated with
        LanguageLinesCount
        top_number (int): number defining how many LanguageLinesCount to return

    Returns:
        list: LanguageLinesCount objects corresponding to the `top_number` most
        present languages
    """
    return sorted(
        project_lc.language_lines_counts, key=lambda x: x.code_count, reverse=True
    )[:top_number]


def top_supported_language_lines_counts(project_lc):
    """Return a list of SupportedLanguage objects corresponding to the supported
    languages detected in the project source archive, sorted by their lines counts.

    Args:
        project_lc (ProjectLinesCount): project lines count object populated with
        LanguageLinesCount

    Returns:
        list: LanguageLinesCount objects corresponding to supported languages
        detected in the project source archive, sorted by their lines counts
    """
    ret = list()
    languages = sorted(
        project_lc.language_lines_counts, key=lambda x: x.code_count, reverse=True
    )
    supported_languages = SupportedLanguage.query.all()
    for c_lang in languages:
        for c_sl in supported_languages:
            if c_sl.name.lower() == c_lang.language.lower():
                ret.append(c_sl)
    return ret


def load_project_lines_count(scc_result):
    """Create a new ProjectLinesCount object and populate it with the given scc
    results.

    Args:
        scc_result (list): deserialized (json.dumps) scc results

    Returns:
        ProjectLinesCount: fully populated project lines count object
    """
    # Empty ProjectLinesCount
    project_lc = ProjectLinesCount(
        total_file_count=0,
        total_line_count=0,
        total_blank_count=0,
        total_comment_count=0,
        total_code_count=0,
        total_complexity_count=0,
    )
    for c in scc_result:
        # Create a LanguageLineCount
        language_lines_count = LanguageLinesCount(
            language=c["Name"],
            file_count=c["Count"],
            line_count=c["Lines"],
            blank_count=c["Blank"],
            comment_count=c["Comment"],
            code_count=c["Code"],
            complexity_count=c["Complexity"],
        )
        project_lc.language_lines_counts.append(language_lines_count)
        # Update ProjectLineCount counters
        project_lc.total_file_count += c["Count"]
        project_lc.total_line_count += c["Lines"]
        project_lc.total_blank_count += c["Blank"]
        project_lc.total_comment_count += c["Comment"]
        project_lc.total_code_count += c["Code"]
        project_lc.total_complexity_count += c["Complexity"]
    return project_lc


def get_user_projects_ids(current_user):
    user_teams = Team.query.filter(
        Team.members.any(username=current_user.username)
    ).all()
    projects_id_list = []

    projects_id_list = [project.id for team in user_teams for project in team.projects]
    projects_id_list = list(set(projects_id_list))
    return projects_id_list


def has_access(current_user, project):
    if current_user.role == ROLE_ADMIN:
        return True
    user_teams = set(
        Team.query.filter(Team.members.any(username=current_user.username)).all()
    )
    project_teams = set(Team.query.filter(Team.projects.any(name=project.name)).all())
    if user_teams.isdisjoint(project_teams):
        return False
    else:
        return True


def generate_xls(project, selected_option):
    wb = Workbook()
    wb.is_spellcheck_enabled = False
    vulnerabilities = project.analysis.vulnerabilities
    # Create the sheets table
    sheets = {}
    # Set the first sheet
    sheets["MainSheet"] = wb.active
    sheets["MainSheet"].title = "MainSheet"
    # Set all vulnerabilitie sheets
    for index, vulnerabilitie in enumerate(vulnerabilities):
        sheet_name = f"vulnerabilitie {vulnerabilitie.id}"
        sheets[sheet_name] = wb.create_sheet(title=sheet_name)
    # Set the MainSheet data
    sheets["MainSheet"].merge_cells("A1:B1")
    sheets["MainSheet"].merge_cells("A2:B2")
    sheets["MainSheet"]["A1"] = "Project name"
    sheets["MainSheet"]["A2"] = project.name
    sheets["MainSheet"]["C1"] = "Project id"
    sheets["MainSheet"]["C2"] = project.id
    sheets["MainSheet"]["D1"] = "Risk level"
    sheets["MainSheet"]["D2"] = project.risk_level
    sheets["MainSheet"]["E1"] = "Lines"
    sheets["MainSheet"]["E2"] = project.project_lines_count.total_line_count
    # Set the vulnerabilites header
    sheets["MainSheet"].row_dimensions[4].height = 40
    sheets["MainSheet"].merge_cells("A4:C4")
    sheets["MainSheet"].merge_cells("G4:J4")
    sheets["MainSheet"].merge_cells("M4:V4")
    for row in sheets["MainSheet"]["A4":"X4"]:
        for cell in row:
            # Center text
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Set cells background color
            cell.fill = PatternFill(
                start_color="DEE6BD", end_color="DEE6BD", fill_type="solid"
            )
    sheets["MainSheet"]["A4"] = "Title"
    sheets["MainSheet"]["D4"] = "Id"
    sheets["MainSheet"]["E4"] = "Confidence"
    sheets["MainSheet"].column_dimensions["E"].width = 15
    sheets["MainSheet"]["F4"] = "Severity"
    sheets["MainSheet"]["G4"] = "Owasp"
    sheets["MainSheet"]["K4"] = "Impact"
    sheets["MainSheet"]["L4"] = "Likelihood"
    sheets["MainSheet"].column_dimensions["L"].width = 15
    sheets["MainSheet"]["M4"] = "Description"
    sheets["MainSheet"]["W4"] = "Occurences"
    sheets["MainSheet"].column_dimensions["W"].width = 15
    sheets["MainSheet"]["X4"] = "Ctrl Click"
    # Take all vulnerabilities
    for index, vulnerabilitie in enumerate(vulnerabilities):
        # Upgrade ["MainSheet"] size
        sheets["MainSheet"].row_dimensions[5 + index].height = 40
        # Set auto line return for description
        sheets["MainSheet"][f"M{5 + index}"].alignment = Alignment(wrap_text=True)
        # Set vulnerabilities data
        sheets["MainSheet"].merge_cells(f"A{5 + index}:C{5 + index}")
        sheets["MainSheet"].merge_cells(f"G{5 + index}:J{5 + index}")
        sheets["MainSheet"].merge_cells(f"M{5 + index}:V{5 + index}")
        sheets["MainSheet"][f"A{5 + index}"] = vulnerabilitie.title
        sheets["MainSheet"][f"D{5 + index}"] = vulnerabilitie.id
        sheets["MainSheet"][f"E{5 + index}"] = vulnerabilitie.confidence
        sheets["MainSheet"][f"F{5 + index}"] = vulnerabilitie.severity
        sheets["MainSheet"][f"G{5 + index}"] = vulnerabilitie.owasp
        sheets["MainSheet"][f"K{5 + index}"] = vulnerabilitie.impact
        sheets["MainSheet"][f"L{5 + index}"] = vulnerabilitie.likelihood
        sheets["MainSheet"][f"M{5 + index}"] = vulnerabilitie.description
        sheets["MainSheet"][f"W{5 + index}"] = len(vulnerabilitie.occurences)
        # Create vulnerabilities occurences sheets button
        sheets["MainSheet"][f"X{5 + index}"] = f"{vulnerabilitie.id} details"
        sheets["MainSheet"][
            f"X{5 + index}"
        ].hyperlink = f"#'vulnerabilitie {vulnerabilitie.id}'!A1"
    # Set vulnerabilities occurences sheets
    for index, vulnerabilitie in enumerate(vulnerabilities):
        # Set mainSheet button
        sheet_name = f"vulnerabilitie {vulnerabilitie.id}"
        sheets[sheet_name].column_dimensions["A"].width = 15
        sheets[sheet_name]["A1"] = "MainSheet"
        # Set sheet data
        sheets[sheet_name]["A1"].hyperlink = f"#'MainSheet'!A1"
        sheets[sheet_name].column_dimensions["B"].width = 15
        sheets[sheet_name]["B1"] = f"Vulnerabilitie {vulnerabilitie.id}"
        sheets[sheet_name].merge_cells("C1:D1")
        sheets[sheet_name].column_dimensions["C"].width = 15
        sheets[sheet_name]["C1"] = vulnerabilitie.title
        # Set occurences header
        sheets[sheet_name].merge_cells("A3:D3")
        sheets[sheet_name]["A3"] = "File path"
        sheets[sheet_name]["E3"] = "Id"
        sheets[sheet_name].merge_cells("F3:Z3")
        sheets[sheet_name]["F3"] = "Match_string"
        for row in sheets[sheet_name]["A3":"Z3"]:
            for cell in row:
                # Center header text
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Set cells background color
                cell.fill = PatternFill(
                    start_color="DEE6BD", end_color="DEE6BD", fill_type="solid"
                )
        # Set occurences data
        for index, occurence in enumerate(vulnerabilitie.occurences):
            if (
                (selected_option == "only_confirmed" and occurence.status == 1)
                or (
                    selected_option == "confirmed_and_undefined"
                    and occurence.status in [1, 0]
                )
                or (selected_option == "all")
            ):
                sheets[sheet_name].merge_cells(f"A{4 + index}:D{4 + index}")
                sheets[sheet_name][f"A{4 + index}"] = occurence.file_path
                sheets[sheet_name][f"E{4 + index}"] = occurence.id
                sheets[sheet_name].merge_cells(f"F{4 + index}:Z{4 + index}")
                sheets[sheet_name][f"F{4 + index}"] = occurence.match_string
    # Save XLS file on disk
    export_folder = os.path.join(
        os.getcwd(), PROJECTS_SRC_PATH, str(project.id), EXPORT_FOLDER_NAME
    )
    if not os.path.isdir(export_folder):
        pathlib.Path(export_folder).mkdir(parents=True, exist_ok=True)
    xls_path = os.path.join(export_folder, project.name + ".xlsx")
    wb.save(xls_path)
    # Return path of the generated XLS file
    return xls_path
